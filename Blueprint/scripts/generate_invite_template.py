from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invite Employees"

    ws["A1"] = "Employee Invite Template"
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Fill rows below, then upload in Manager portal → Invite employees from Excel."
    ws.merge_cells("A2:C2")
    ws["A2"].font = Font(size=10, italic=True, color="4B5563")

    # Backend accepts: Email, Name; Department is required only when HR uploads.
    headers = ["Email", "Name", "Department"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0F766E")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Example rows
    examples = [
        ("alice@yourcompany.com", "Alice Sharma", "Software"),
        ("bob@yourcompany.com", "Bob Singh", "Software"),
    ]
    for i, row in enumerate(examples, start=1):
        for col, v in enumerate(row, start=1):
            ws.cell(row=header_row + i, column=col, value=v)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:C{header_row}"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 20

    # Data validations (lightweight)
    email_dv = DataValidation(type="custom", formula1='ISNUMBER(SEARCH("@",A5))', allow_blank=False)
    email_dv.errorTitle = "Invalid email"
    email_dv.error = "Enter a valid email address"
    ws.add_data_validation(email_dv)
    email_dv.add("A5:A500")

    name_dv = DataValidation(type="custom", formula1="LEN(TRIM(B5))>0", allow_blank=False)
    name_dv.errorTitle = "Missing name"
    name_dv.error = "Name is required"
    ws.add_data_validation(name_dv)
    name_dv.add("B5:B500")

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Column rules"
    notes["A1"].font = Font(bold=True, size=12)
    notes["A3"] = "Email"
    notes["B3"] = "Required. Must be a company email ending with the same domain as the logged-in Manager/HR."
    notes["A4"] = "Name"
    notes["B4"] = "Required. Employee full name."
    notes["A5"] = "Department"
    notes["B5"] = "Required ONLY when HR uploads. For Managers, department is taken from the manager profile."
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 90

    out = "Employee_Invite_Template.xlsx"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()

